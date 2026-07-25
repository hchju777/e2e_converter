import tempfile
import unittest
from pathlib import Path

import openpyxl

from src.utils.banner import REQUIRED_DASHBOARD_BANNERS
from src.utils.excel_history import (
    METRICS_PER_BANNER,
    ExcelHistoryError,
    check_new_wave,
    next_wave,
    normalize_wave,
    parse_wave,
    read_history,
    read_wave_labels,
    validate_against_spec,
    wave_start_date,
    write_history,
)


def build_workbook(path: Path, waves: list[str], filled: int | None = None) -> Path:
    """테스트용 Summary 시트를 만든다. filled 이후의 차수 열은 헤더만 두고 값을 비운다."""
    filled = len(waves) if filled is None else filled
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["구분", "항목", "문항", "SAV 변수", "계산", "배너조건", *waves])
    for banner in REQUIRED_DASHBOARD_BANNERS:
        for index in range(METRICS_PER_BANNER):
            values = [
                float(index + position) if position < filled else None
                for position in range(len(waves))
            ]
            sheet.append(["전체", f"항목{index}", f"문항{index}", "PID", "Count", banner, *values])
    workbook.save(path)
    return path


class WaveParsingTests(unittest.TestCase):
    def test_parses_and_normalizes(self):
        self.assertEqual(parse_wave("26년 6차"), (26, 6))
        self.assertEqual(normalize_wave("26년6차"), "26년 6차")

    def test_rejects_bad_format_and_range(self):
        with self.assertRaisesRegex(ExcelHistoryError, "형식"):
            parse_wave("2026-06")
        with self.assertRaisesRegex(ExcelHistoryError, "지원하지 않는"):
            parse_wave("26년 13차")


class ReadHistoryTests(unittest.TestCase):
    def setUp(self):
        self.dir = Path(tempfile.mkdtemp())

    def test_reads_only_filled_waves(self):
        path = build_workbook(self.dir / "h.xlsx", ["25년 1차", "25년 2차", "25년 3차"], filled=2)
        history = read_history(path)

        self.assertEqual(history.waves, ["25년 1차", "25년 2차"])  # 빈 열은 건너뛴다
        self.assertEqual(history.latest_wave, "25년 2차")
        self.assertEqual(len(history.values["25년 1차"]), len(REQUIRED_DASHBOARD_BANNERS))
        self.assertEqual(len(history.values["25년 1차"]["T3"]), METRICS_PER_BANNER)

    def test_rejects_missing_summary_sheet(self):
        workbook = openpyxl.Workbook()
        workbook.active.title = "다른시트"
        path = self.dir / "bad.xlsx"
        workbook.save(path)

        with self.assertRaisesRegex(ExcelHistoryError, "Summary"):
            read_history(path)

    def test_rejects_out_of_order_waves(self):
        path = build_workbook(self.dir / "order.xlsx", ["25년 3차", "25년 1차"])
        with self.assertRaisesRegex(ExcelHistoryError, "시간 순서"):
            read_history(path)

    def test_rejects_when_no_wave_has_values(self):
        path = build_workbook(self.dir / "empty.xlsx", ["25년 1차"], filled=0)
        with self.assertRaisesRegex(ExcelHistoryError, "값이 채워진 차수가"):
            read_history(path)


class WaveLabelTests(unittest.TestCase):
    def setUp(self):
        self.dir = Path(tempfile.mkdtemp())

    def test_lists_only_filled_waves(self):
        path = build_workbook(self.dir / "h.xlsx", ["25년 1차", "25년 2차", "25년 3차"], filled=2)
        self.assertEqual(read_wave_labels(path), ["25년 1차", "25년 2차"])

    def test_start_date_is_first_day_of_that_month(self):
        self.assertEqual(wave_start_date("25년 1차"), "2025-01-01")
        self.assertEqual(wave_start_date("26년 12차"), "2026-12-01")

    def test_next_wave_follows_the_last_one(self):
        self.assertEqual(next_wave("26년 5차"), "26년 6차")
        self.assertEqual(next_wave("26년 11차"), "26년 12차")

    def test_next_wave_rolls_over_the_year(self):
        self.assertEqual(next_wave("26년 12차"), "27년 1차")
        self.assertEqual(next_wave("99년 12차"), "00년 1차")

    def test_next_wave_is_always_a_valid_wave(self):
        for wave in ("25년 1차", "26년 12차", "99년 12차"):
            parse_wave(next_wave(wave))  # 형식이 깨지면 여기서 실패한다

    def test_rejects_workbook_without_summary(self):
        workbook = openpyxl.Workbook()
        workbook.active.title = "다른시트"
        path = self.dir / "bad.xlsx"
        workbook.save(path)
        with self.assertRaisesRegex(ExcelHistoryError, "Summary"):
            read_wave_labels(path)

    def test_suggests_next_year_when_workbook_ends_in_december(self):
        """마지막이 12차인 엑셀을 올리면 다음 해 1차를 제안해야 한다."""
        waves = [f"26년 {n}차" for n in range(1, 13)]
        path = build_workbook(self.dir / "full.xlsx", waves)

        labels = read_wave_labels(path)
        self.assertEqual(labels[-1], "26년 12차")
        self.assertEqual(next_wave(labels[-1]), "27년 1차")


class SpecAlignmentTests(unittest.TestCase):
    def setUp(self):
        self.dir = Path(tempfile.mkdtemp())
        self.history = read_history(build_workbook(self.dir / "h.xlsx", ["25년 1차"]))

    def test_accepts_matching_spec(self):
        validate_against_spec(self.history, [f"문항{i}" for i in range(METRICS_PER_BANNER)])

    def test_rejects_misaligned_spec(self):
        questions = [f"문항{i}" for i in range(METRICS_PER_BANNER)]
        questions[7] = "다른문항"
        with self.assertRaisesRegex(ExcelHistoryError, "순서가 metric_spec"):
            validate_against_spec(self.history, questions)


class NewWaveTests(unittest.TestCase):
    def setUp(self):
        self.dir = Path(tempfile.mkdtemp())
        self.history = read_history(
            build_workbook(self.dir / "h.xlsx", ["25년 1차", "25년 2차"])
        )

    def test_accepts_next_wave(self):
        self.assertEqual(check_new_wave(self.history, "25년3차"), "25년 3차")

    def test_rejects_duplicate_wave(self):
        with self.assertRaisesRegex(ExcelHistoryError, "이미 존재하는 차수"):
            check_new_wave(self.history, "25년 2차")

    def test_rejects_earlier_wave(self):
        history = read_history(build_workbook(self.dir / "h2.xlsx", ["25년 5차"]))
        with self.assertRaisesRegex(ExcelHistoryError, "앞섭니다"):
            check_new_wave(history, "25년 3차")


class WriteHistoryTests(unittest.TestCase):
    def setUp(self):
        self.dir = Path(tempfile.mkdtemp())
        self.values = {
            banner: [float(i) for i in range(METRICS_PER_BANNER)]
            for banner in REQUIRED_DASHBOARD_BANNERS
        }

    def test_round_trip_preserves_past_and_adds_new(self):
        source = build_workbook(self.dir / "h.xlsx", ["25년 1차", "25년 2차"])
        before = read_history(source)

        output = write_history(source, self.dir / "out.xlsx", "25년 3차", self.values)
        after = read_history(output)

        self.assertEqual(after.waves, ["25년 1차", "25년 2차", "25년 3차"])
        for wave in before.waves:
            self.assertEqual(after.values[wave]["전체"], before.values[wave]["전체"])
        self.assertEqual(after.values["25년 3차"]["T2"], self.values["T2"])
        self.assertNotEqual(source.read_bytes(), output.read_bytes())  # 원본은 그대로 둔다

    def test_reuses_preexisting_empty_column(self):
        source = build_workbook(self.dir / "h.xlsx", ["25년 1차", "25년 2차"], filled=1)
        output = write_history(source, self.dir / "out.xlsx", "25년 2차", self.values)

        sheet = openpyxl.load_workbook(output, read_only=True)["Summary"]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        self.assertEqual(headers.count("25년 2차"), 1)  # 새 열을 만들지 않고 빈 열을 채운다
        self.assertEqual(read_history(output).waves, ["25년 1차", "25년 2차"])

    def test_rejects_overwriting_filled_wave(self):
        source = build_workbook(self.dir / "h.xlsx", ["25년 1차"])
        with self.assertRaisesRegex(ExcelHistoryError, "이미 값이 있는 차수"):
            write_history(source, self.dir / "out.xlsx", "25년 1차", self.values)

    def test_rejects_wrong_value_count(self):
        source = build_workbook(self.dir / "h.xlsx", ["25년 1차"])
        broken = dict(self.values)
        broken["T1"] = [1.0, 2.0]
        with self.assertRaisesRegex(ExcelHistoryError, "329개여야"):
            write_history(source, self.dir / "out.xlsx", "25년 2차", broken)
