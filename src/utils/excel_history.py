"""히스토릭 지표 엑셀(Summary 시트)을 읽고 쓴다.

이 엑셀이 조사 차수별 지표 값의 정본이다. 대시보드 HTML에는 값을 하드코딩하지 않고,
매번 이 엑셀의 모든 차수 + 새로 계산한 SAV 차수로 다시 채운다.

Summary 시트 구조 (0-indexed 행/열 기준):
    행 0            헤더. 0~5열은 지표 정의(구분/항목/문항/SAV 변수/계산/배너조건),
                    6열부터 차수별 값이 한 차수당 한 열씩 들어간다.
    행 1~329        '전체' 배너 블록 (metric_spec.csv와 같은 순서의 329개 지표)
    행 330~658      '수도권' 블록, 이후 지방·T1·T2·T3가 같은 크기로 이어진다.

차수 열은 미리 만들어 두고 값을 비워 둘 수 있다(예: 26년 7~12차). 값이 비어 있는
열은 아직 조사하지 않은 차수이므로 읽을 때 건너뛴다.
"""

import re
import shutil
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from src.utils.banner import REQUIRED_DASHBOARD_BANNERS
from src.utils.logger import get_logger


logger = get_logger(__name__)

SHEET_NAME = "Summary"
SPEC_COLUMN_COUNT = 6  # 구분/항목/문항/SAV 변수/계산/배너조건
METRICS_PER_BANNER = 329
QUESTION_COLUMN = 2  # 0-indexed. '문항'

_WAVE_RE = re.compile(r"^\s*(\d{2})년\s*(\d{1,2})차\s*$")


class ExcelHistoryError(ValueError):
    """히스토릭 엑셀의 구조나 내용이 기대와 다를 때 발생한다."""


def parse_wave(wave: str) -> tuple[int, int]:
    """'26년 6차' -> (26, 6). 정렬과 순서 검증에 쓴다."""
    match = _WAVE_RE.match(str(wave))
    if not match:
        raise ExcelHistoryError(f"차수 형식이 올바르지 않습니다(예: 26년 6차): {wave}")
    year, number = int(match.group(1)), int(match.group(2))
    if not 1 <= number <= 12:
        raise ExcelHistoryError(f"지원하지 않는 차수입니다: {wave}")
    return year, number


def normalize_wave(wave: str) -> str:
    """공백 차이를 없앤 표준 표기('26년 6차')로 바꾼다."""
    year, number = parse_wave(wave)
    return f"{year}년 {number}차"


def _banner_block_start(index: int) -> int:
    """0-indexed 배너 순번 -> Summary 시트에서 그 블록이 시작하는 0-indexed 행."""
    return 1 + index * METRICS_PER_BANNER


@dataclass
class ExcelHistory:
    """엑셀에서 읽어온 차수별 지표 값."""

    waves: list[str]  # 값이 채워진 차수 (엑셀 열 순서)
    values: dict[str, dict[str, list]]  # {차수: {배너: 329개 값}}
    questions: list[str]  # '전체' 블록의 문항 329개 (스펙 정합성 확인용)
    source_path: Path

    @property
    def latest_wave(self) -> str | None:
        return self.waves[-1] if self.waves else None


def read_history(path: str | Path) -> ExcelHistory:
    """Summary 시트에서 값이 채워진 차수를 모두 읽는다."""
    source = Path(path)
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ExcelHistoryError(
                f"엑셀에서 '{SHEET_NAME}' 시트를 찾을 수 없습니다: {source.name}"
            )
        sheet = workbook[SHEET_NAME]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    expected_rows = 1 + len(REQUIRED_DASHBOARD_BANNERS) * METRICS_PER_BANNER
    if len(rows) < expected_rows:
        raise ExcelHistoryError(
            f"Summary 시트의 행이 부족합니다. {expected_rows}행이 필요한데 {len(rows)}행입니다. "
            "배너 6개 × 지표 329개 구조인지 확인해 주세요."
        )

    header = rows[0]
    wave_columns: list[tuple[str, int]] = []
    for column in range(SPEC_COLUMN_COUNT, len(header)):
        label = header[column]
        if label is None or not _WAVE_RE.match(str(label)):
            continue
        # 값이 비어 있는 열은 아직 조사하지 않은 차수이므로 건너뛴다.
        if rows[1][column] is None:
            continue
        wave_columns.append((normalize_wave(str(label)), column))

    if not wave_columns:
        raise ExcelHistoryError("엑셀에 값이 채워진 차수가 하나도 없습니다.")

    duplicates = {w for w, _ in wave_columns if [x for x, _ in wave_columns].count(w) > 1}
    if duplicates:
        raise ExcelHistoryError(f"엑셀에 중복된 차수 열이 있습니다: {sorted(duplicates)}")

    keys = [parse_wave(w) for w, _ in wave_columns]
    if keys != sorted(keys):
        raise ExcelHistoryError(
            f"엑셀의 차수 열이 시간 순서가 아닙니다: {[w for w, _ in wave_columns]}"
        )

    values: dict[str, dict[str, list]] = {}
    for wave, column in wave_columns:
        per_banner: dict[str, list] = {}
        for index, banner in enumerate(REQUIRED_DASHBOARD_BANNERS):
            start = _banner_block_start(index)
            per_banner[banner] = [
                _cell_number(rows[start + offset][column]) for offset in range(METRICS_PER_BANNER)
            ]
        values[wave] = per_banner

    questions = [
        str(rows[_banner_block_start(0) + offset][QUESTION_COLUMN]).strip()
        for offset in range(METRICS_PER_BANNER)
    ]

    logger.info(
        "📗 히스토릭 엑셀 로드: %s (차수 %d개: %s ~ %s)",
        source.name, len(wave_columns), wave_columns[0][0], wave_columns[-1][0],
    )
    return ExcelHistory(
        waves=[w for w, _ in wave_columns],
        values=values,
        questions=questions,
        source_path=source,
    )


def read_wave_labels(path: str | Path) -> list[str]:
    """값이 채워진 차수 목록만 가볍게 읽는다(지표 값은 읽지 않는다).

    업로드 직후 화면에 조사 기간 기본값을 채워 넣는 용도라 전체를 읽을 필요가 없다.
    """
    workbook = openpyxl.load_workbook(Path(path), data_only=True, read_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ExcelHistoryError(f"엑셀에서 '{SHEET_NAME}' 시트를 찾을 수 없습니다.")
        rows = workbook[SHEET_NAME].iter_rows(min_row=1, max_row=2, values_only=True)
        header = next(rows, None)
        first_row = next(rows, None)
    finally:
        workbook.close()

    if header is None or first_row is None:
        raise ExcelHistoryError("Summary 시트가 비어 있습니다.")

    waves = []
    for column in range(SPEC_COLUMN_COUNT, len(header)):
        label = header[column]
        if label is None or not _WAVE_RE.match(str(label)):
            continue
        if column >= len(first_row) or first_row[column] is None:
            continue  # 값이 비어 있는 열은 아직 조사하지 않은 차수다.
        waves.append(normalize_wave(str(label)))
    if not waves:
        raise ExcelHistoryError("엑셀에 값이 채워진 차수가 하나도 없습니다.")
    return waves


def wave_start_date(wave: str) -> str:
    """차수를 그 달 1일의 'YYYY-MM-DD'로 바꾼다 (25년 1차 -> 2025-01-01)."""
    year, number = parse_wave(wave)
    return f"20{year:02d}-{number:02d}-01"


def next_wave(wave: str) -> str:
    """바로 다음 차수를 돌려준다 (26년 5차 -> 26년 6차, 26년 12차 -> 27년 1차)."""
    year, number = parse_wave(wave)
    if number >= 12:
        return f"{(year + 1) % 100:02d}년 1차"
    return f"{year}년 {number + 1}차"


def _cell_number(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def validate_against_spec(history: ExcelHistory, spec_questions: list[str]) -> None:
    """엑셀의 지표 행 순서가 metric_spec.csv와 같은지 확인한다.

    행이 어긋나면 값이 조용히 잘못된 지표에 들어가므로 반드시 먼저 막는다.
    """
    if len(spec_questions) != METRICS_PER_BANNER:
        raise ExcelHistoryError(
            f"지표 스펙이 {METRICS_PER_BANNER}개여야 하는데 {len(spec_questions)}개입니다."
        )
    for index, (excel_question, spec_question) in enumerate(zip(history.questions, spec_questions)):
        if str(excel_question).strip() != str(spec_question).strip():
            raise ExcelHistoryError(
                f"엑셀의 지표 순서가 metric_spec.csv와 다릅니다 "
                f"({index + 1}번째 문항: 엑셀 '{excel_question}' vs 스펙 '{spec_question}'). "
                "다른 버전의 엑셀을 올렸는지 확인해 주세요."
            )


def check_new_wave(history: ExcelHistory, wave: str) -> str:
    """새로 추가할 차수가 유효한지 확인하고 표준 표기를 돌려준다."""
    new_wave = normalize_wave(wave)
    if new_wave in history.waves:
        raise ExcelHistoryError(
            f"엑셀에 이미 존재하는 차수입니다: {new_wave}. "
            "다른 차수를 입력하거나 이전 차수가 없는 엑셀을 사용해 주세요."
        )
    latest = history.latest_wave
    if latest and parse_wave(new_wave) < parse_wave(latest):
        raise ExcelHistoryError(
            f"새 차수({new_wave})가 엑셀의 마지막 차수({latest})보다 앞섭니다. "
            "차수를 다시 확인해 주세요."
        )
    return new_wave


def write_history(
    source_path: str | Path,
    output_path: str | Path,
    wave: str,
    banner_values: dict[str, list],
) -> Path:
    """원본 엑셀을 복사해 새 차수 열을 채워 넣는다.

    비어 있는 채로 미리 만들어 둔 차수 열이 있으면 그 자리에 쓰고, 없으면 맨 뒤에 새 열을 만든다.
    """
    new_wave = normalize_wave(wave)
    source, output = Path(source_path), Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if source.resolve() != output.resolve():
        shutil.copyfile(source, output)

    workbook = openpyxl.load_workbook(output)
    try:
        sheet = workbook[SHEET_NAME]
        target = _target_column(sheet, new_wave)
        sheet.cell(row=1, column=target, value=new_wave)
        for index, banner in enumerate(REQUIRED_DASHBOARD_BANNERS):
            start = _banner_block_start(index)
            values = banner_values[banner]
            if len(values) != METRICS_PER_BANNER:
                raise ExcelHistoryError(
                    f"'{banner}' 배너 값이 {METRICS_PER_BANNER}개여야 하는데 {len(values)}개입니다."
                )
            for offset, value in enumerate(values):
                # openpyxl은 1-indexed이므로 0-indexed 행에 1을 더한다.
                sheet.cell(row=start + offset + 1, column=target, value=value)
        workbook.save(output)
    finally:
        workbook.close()

    logger.info("💾 히스토릭 엑셀 저장: %s (%s 추가)", output.name, new_wave)
    return output


def _target_column(sheet, wave: str) -> int:
    """새 차수를 쓸 1-indexed 열 번호를 찾는다."""
    last_used = SPEC_COLUMN_COUNT  # 1-indexed로 F열
    for column in range(SPEC_COLUMN_COUNT + 1, sheet.max_column + 1):
        label = sheet.cell(row=1, column=column).value
        if label is None:
            continue
        last_used = column
        if _WAVE_RE.match(str(label)) and normalize_wave(str(label)) == wave:
            if sheet.cell(row=2, column=column).value is not None:
                raise ExcelHistoryError(f"엑셀에 이미 값이 있는 차수입니다: {wave}")
            return column  # 미리 만들어 둔 빈 열을 재사용한다.
    return last_used + 1
